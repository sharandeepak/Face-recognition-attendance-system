import xlsxwriter
import xlrd
import cv2
import os
import numpy as np
import faceRecognition as fr
import urllib.request
import time
from win10toast import ToastNotifier
import pyqrcode
from PIL import Image
import cv2
import numpy as np
import pyzbar.pyzbar as pyzbar
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from datetime import datetime
from gtts import gTTS
from playsound import playsound
import openpyxl
import sqlite3


# faces,faceID=fr.labels_for_training_data('C:/Users/sharan/Desktop/project_try3/training_images')
# face_recognizer=fr.train_classifier(faces,faceID)
# face_recognizer.save('trainingData.yml')


def qr():
    cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)
    # address="http://192.168.43.1:8080/video"
    # cap.open(address)
    font = cv2.FONT_HERSHEY_PLAIN

    while True:
        _, frame = cap.read()

        decodedObjects = pyzbar.decode(frame)
        for obj in decodedObjects:
            #print("Data", obj.data)
            print(decodedObjects[0].data.decode('ascii'))
            cv2.putText(frame, decodedObjects[0].data.decode(
                'ascii'), (50, 50), font, 2, (255, 0, 0), 3)
            if(decodedObjects[0].type == "qrcode" or decodedObjects[0].type == "QRCODE"):
                valuee = decodedObjects[0].data.decode('ascii')
                leng = len(valuee)
                dotpos = valuee.find('.')
                personcode = valuee[0:dotpos]
                namee = valuee[dotpos+1:leng]
                print("personcode  "+personcode)

                if personcode.isnumeric():
                    print("hello")
                    imgg = cv2.imread(
                        "training_images/"+str(personcode)+"/"+str(personcode)+".10.jpg")
                    cv2.imshow("face", imgg)
                    fr.put_text(imgg, personcode, 0, 0)
                    cv2.moveWindow("face", 0, 0)
                    playsound("qrdone.wav")
                    print("done qr")
                    facee(personcode)

        resized_img = cv2.resize(frame, (800, 800))
        cv2.imshow("Frame", resized_img)
        cv2.moveWindow("Frame", 0, 0)

        if cv2.waitKey(100) & 0xFF == ord('q'):
            cap.release()
            cv2.destroyAllWindows()
            break
        if cv2.getWindowProperty('Frame', cv2.WND_PROP_VISIBLE) < 1:
            cap.release()
            cv2.destroyAllWindows()
            break


def facee(codee):
    print(type(codee))
    cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)
    print("beginning of facee")
    face_recognizer = cv2.face.LBPHFaceRecognizer_create()
    face_recognizer.read("trainingData.yml")
    print("after reading")
    # url="http://192.168.43.1:8080/shot.jpg"

    # cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)

    # address="http://192.168.43.1:8080/video"
    # cap.open(address)

    while True:
        ret, test_img = cap.read()
        # imgResp=urllib.request.urlopen(url)
        # imgNp=np.array(bytearray(imgResp.read()),dtype=np.uint8)
        # test_img=cv2.imdecode(imgNp,-1)
        faces_detected, gray_img = fr.faceDetection(test_img)

        for (x, y, w, h) in faces_detected:
            cv2.rectangle(test_img, (x, y), (x+w, y+h),
                          (255, 0, 0), thickness=3)

        resized_img = cv2.resize(test_img, (1000, 600))
        cv2.imshow("face detection", resized_img)
        cv2.moveWindow("face detection", 700, 100)

        cv2.waitKey(10)

        for face in faces_detected:
            (x, y, w, h) = face
            roi_gray = gray_img[y:y+h, x:x+w]
            label, confidence = face_recognizer.predict(roi_gray)
            print("confidence:", confidence)
            print("label:", label)
            fr.draw_rect(test_img, face)
            if str(label) == str(codee):
                print("YESS")
            else:
                print("No"+"Label"+str(label)+"codee"+str(codee))
            if confidence < 50 and str(label) == str(codee):
                # toaster=ToastNotifier()
                # toaster.show_toast("hello",predicted_name,"you have marked present")
                writeTodb(codee)
                qr()

        resized_img = cv2.resize(test_img, (1000, 600))
        cv2.imshow("face detection", resized_img)
        # print("confidence:",confidence)

        if cv2.waitKey(100) & 0xFF == ord('q'):
            cap.release()
            cv2.destroyAllWindows()
            break
        if cv2.getWindowProperty('face detection', cv2.WND_PROP_VISIBLE) < 1:
            cap.release()
            cv2.destroyAllWindows()
            break


def writeAttendance(roww, coll, timee):
    # ONLINE -------------------------------
    # use creds to create a client to interact with the Google Drive API
    scope = ['https://spreadsheets.google.com/feeds',
             'https://www.googleapis.com/auth/drive']

    creds = ServiceAccountCredentials.from_json_keyfile_name(
        'client.json', scope)
    client = gspread.authorize(creds)

    # Find a workbook by name and open the first sheet
    # Make sure you use the right name here.
    sheet = client.open("Attendance Report").sheet1

    if sheet.cell(roww, coll-1).value == '':
        sheet.update_cell(roww, coll-1, timee)
        # output=gTTS(text="hello "+nameqr,lang="en-in",slow=False)
        # output.save("output.mp3")

    else:
        sheet.update_cell(roww, coll, timee)
        sheet.update_cell(roww, coll+1, "P")
        val = sheet.cell(roww, 99).value
        if val == '':
            sheet.update_cell(roww, 99, 1)
        else:
            sheet.update_cell(roww, 99, int(val)+1)

    # OFFLINE RECORDS----------------------

    # path = "Attendance Report/Attendance Report.xlsx"
    # excel_workbook = openpyxl.load_workbook(path)
    # excel_sheet = excel_workbook['Sheet']

    # if excel_sheet.cell(row=int(roww), column=int(coll)).value == None:
    #     now = datetime.now()
    #     time = now.strftime("%H:%M")
    #     print("row:"+str(roww)+"  col="+str(coll))
    #     excel_sheet.cell(row=int(roww), column=int(coll)).value = str(time)
    #     excel_workbook.save(path)
    #     # output = gTTS(text="hello "+nameqr, lang="en-in", slow=False)
    #     # output.save("output.mp3")
    #     playsound("thankyou.mp3")
    # else:
    #     if excel_sheet.cell(row=int(roww), column=int(coll+1)).value == None:
    #         now = datetime.now()
    #         time = now.strftime("%H:%M")
    #         excel_sheet.cell(row=int(roww), column=int(coll+1)).value = time
    #         excel_sheet.cell(row=int(roww), column=int(coll+2)).value = "P"
    #         val = excel_sheet.cell(row=int(roww), column=int(99)).value
    #         if val == None:
    #             excel_sheet.cell(row=int(roww), column=int(99)).value = 1
    #         else:
    #             excel_sheet.cell(row=int(roww), column=int(99)
    #                              ).value = int(val)+1
    #         excel_workbook.save(path)
    #         playsound("thankyou.mp3")


def excelWrite(codee, datee, timee):

    # ---------------ONLINE-----------------------------

    # use creds to create a client to interact with the Google Drive API
    scope = ['https://spreadsheets.google.com/feeds',
             'https://www.googleapis.com/auth/drive']

    creds = ServiceAccountCredentials.from_json_keyfile_name(
        'client.json', scope)
    client = gspread.authorize(creds)

    sheet = client.open("Attendance Report").sheet1

    findDate = sheet.row_values(1)
    findUIN = sheet.col_values(2)
    coll = findDate.index(str(datee))+1
    roww = findUIN.index(codee)+1
    writeAttendance(roww, coll, timee)

    # --------------OFFLINE---------------------

    # now = datetime.now()
    # date = now.strftime("%d/%m")
    # path = "Attendance Report/Attendance Report.xlsx"
    # excel_workbook = xlrd.open_workbook(path)
    # excel_sheet = excel_workbook.sheet_by_index(0)
    # findDate = excel_sheet.row_values(0)
    # coll = findDate.index(str(date))
    # findName = excel_sheet.col_values(0)
    # roww = findName.index(int(codee))+1
    # print("row="+str(roww)+"  column="+str(coll))
    # writeAttendance(roww, coll)


def writeTodb(uid_db):
    connection = sqlite3.connect('attendance.db')
    cursor = connection.cursor()
    try:
        cursor.execute(
            'CREATE TABLE MarkAttendance(id INTEGER PRIMARY KEY,UID INTEGER,datee TEXT,Timee TEXT)')
    except:
        print("File Already Exist")
    finally:
        now = datetime.now()
        date = now.strftime("%d/%m")
        time = now.strftime("%H:%M")
        cursor.execute(
            'INSERT INTO MarkAttendance(UID,datee,Timee) VALUES(?,?,?)', (uid_db, date, time))
        connection.commit()
        playsound("thankyou.mp3")


def readFromdb():
    connection = sqlite3.connect('attendance.db')
    cursor = connection.cursor()
    cursor.execute('SELECT * FROM MarkAttendance')
    row = cursor.fetchone()
    while row is not None:
        id = row[0]
        uin = row[1]
        datee = row[2]
        Time2 = row[3]
        # call fun(uin,inTime,outTime)
        print("\nid: "+str(id)+"\nUIN : "+str(uin)+"\ndatee : "+datee+"\nTime : "+str(Time2))
        excelWrite(str(uin), datee, Time2)
        # print("SUCCESSFULLY DELETED")
        row = cursor.fetchone()
        connection.commit()
    # deleteFromdb(id)
    connection.commit()
    return 2
def deleteFromdb(id):
    connection = sqlite3.connect('attendance.db')
    cursor = connection.cursor()
    cursor.execute("DELETE FROM MarkAttendance WHERE ID=?",(id,))
    connection.commit()

